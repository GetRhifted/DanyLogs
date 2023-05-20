from django.shortcuts import render, redirect
from django.urls import reverse_lazy, reverse
from django.http import HttpResponse
from django.views.generic import ListView, TemplateView, DetailView, UpdateView, FormView, DeleteView, CreateView
from django.contrib.auth.views import LoginView, LogoutView
from django.db.models import Q
from django.conf import settings
from django.core.exceptions import ValidationError
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib import messages

from .forms import RegistroGeneralForm, RegistroIngredientesForm, RegistroTiemposForm, RegistroDesechosForm, RegistroBrixForm, RegistroPaquetesForm, CompararRegistrosForm, RegistrodeUsuarioForm
from. models import Registro

import openpyxl
from openpyxl.styles import Border, Side, PatternFill
from formtools.wizard.views import SessionWizardView



class HomeView(ListView):
    model = Registro
    template_name = 'registros/home.html'
    queryset = Registro.objects.order_by('-id')[:10]

class RegistroWizardView(SessionWizardView):
    template_name = 'registros/registro_wizard.html'
    form_list = [RegistroGeneralForm, RegistroIngredientesForm, RegistroTiemposForm, RegistroDesechosForm,
                 RegistroBrixForm, RegistroPaquetesForm]
    #success_url = reverse_lazy('registros:registro_completado')

    def done(self, form_list, **kwargs):
        registro = Registro(Bache=form_list[0].cleaned_data['Bache'],
                            Fecha=form_list[0].cleaned_data['Fecha'],
                            Gramos_de_Mora=form_list[1].cleaned_data['Gramos_de_Mora'],
                            Gramos_de_Azucar=form_list[1].cleaned_data['Gramos_de_Azucar'],
                            Gramos_de_Sorbato=form_list[1].cleaned_data['Gramos_de_Sorbato'],
                            Hora_Inicio=form_list[2].cleaned_data['Hora_Inicio'],
                            Primer_Hervor=form_list[2].cleaned_data['Primer_Hervor'],
                            Pausa_de_enfriado=form_list[2].cleaned_data['Pausa_de_enfriado'],
                            Despulpado=form_list[2].cleaned_data['Despulpado'],
                            Ultima_Coccion=form_list[2].cleaned_data['Ultima_Coccion'],
                            Hora_Final=form_list[2].cleaned_data['Hora_Final'],
                            Desechos_Mora=form_list[3].cleaned_data['Desechos_Mora'],
                            Semilla=form_list[3].cleaned_data['Semilla'],
                            Pulpa=form_list[3].cleaned_data['Pulpa'],
                            Valor_Primer_Brix=form_list[4].cleaned_data['Valor_Primer_Brix'],
                            Hora_Primer_Brix=form_list[4].cleaned_data['Hora_Primer_Brix'],
                            Valor_Brix_Final=form_list[4].cleaned_data['Valor_Brix_Final'],
                            Hora_Brix_Final=form_list[4].cleaned_data['Hora_Brix_Final'],
                            Paquete_250_gr=form_list[5].cleaned_data['Paquete_250_gr'],
                            Paquete_500_gr=form_list[5].cleaned_data['Paquete_500_gr'],
                            Paquete_5000_gr=form_list[5].cleaned_data['Paquete_5000_gr'],
                            )
        registro.save()
        return redirect('registros:registro_completado')
    
    def get_success_url(self):
        return reverse('registros:registro_completado')
        
class RegistroCompletadoView(TemplateView):
    model = Registro
    form_class = RegistroGeneralForm
    template_name = 'registros/registro_completado.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['registro'] = Registro.objects.last()
        return context

    def form_valid(self, form):
        # Guarda el registro y asigna la instancia del objeto a la variable 'registro'
        registro = form.save()
        # Agrega el objeto de registro al contexto de la plantilla
        return render(self.request, self.template_name, {'registro': registro})
    
class RegistroDetalladoView(DetailView):
    model = Registro
    template_name = 'registros/registro_detallado.html'

class RegistroUpdateView(UpdateView):
     model = Registro
     form_class = RegistroGeneralForm
     template_name = 'registros/edicion_registro.html'
     
     def get_success_url(self):
        return reverse_lazy('registros:edicion_registro2', kwargs={'pk': self.object.pk})

class RegistroUpdateView2(UpdateView):
     model = Registro
     form_class = RegistroIngredientesForm
     template_name = 'registros/edicion_registro2.html'
     
     def get_success_url(self):
        return reverse_lazy('registros:edicion_registro3', kwargs={'pk': self.object.pk})

class RegistroUpdateView3(UpdateView):
     model = Registro
     form_class = RegistroTiemposForm
     template_name = 'registros/edicion_registro3.html'
     
     def get_success_url(self):
        return reverse_lazy('registros:edicion_registro4', kwargs={'pk': self.object.pk})

class RegistroUpdateView4(UpdateView):
     model = Registro
     form_class = RegistroDesechosForm
     template_name = 'registros/edicion_registro4.html'
     
     def get_success_url(self):
        return reverse_lazy('registros:edicion_registro5', kwargs={'pk': self.object.pk})

class RegistroUpdateView5(UpdateView):
     model = Registro
     form_class = RegistroBrixForm
     template_name = 'registros/edicion_registro5.html'
     
     def get_success_url(self):
        return reverse_lazy('registros:edicion_registro6', kwargs={'pk': self.object.pk})

class RegistroUpdateView6(UpdateView):
     model = Registro
     form_class = RegistroPaquetesForm
     template_name = 'registros/edicion_registro6.html'
     
     def get_success_url(self):
        return reverse_lazy('registros:registro_detallado', kwargs={'pk': self.object.pk})


class RegistroDeleteView(DeleteView):
    model = Registro
    template_name = 'registros/borrar_registro.html'
    success_url = reverse_lazy('registros:home')


class CompararRegistrosView(FormView):
    template_name = 'registros/comparar_registros.html'
    form_class = CompararRegistrosForm

    def form_valid(self, form):
        registros = form.cleaned_data['registros']
        if registros.count() < settings.NUMERO_MINIMO_REGISTROS:
            form.add_error(
                'registros',
                ValidationError(f'Seleccione al menos {settings.NUMERO_MINIMO_REGISTROS} registros.'),
            )
            return self.form_invalid(form)

        # Obtener los datos de los registros seleccionados
        datos = []
        for registro in registros:
            datos.append([
                registro.Bache,
                registro.Fecha,
                registro.Gramos_de_Mora,
                registro.Gramos_de_Azucar,
                registro.Gramos_de_Sorbato,
                registro.Hora_Inicio,
                registro.Primer_Hervor,
                registro.Pausa_de_enfriado,
                registro.Despulpado,
                registro.Ultima_Coccion,
                registro.Hora_Final,
                registro.Desechos_Mora,
                registro.Semilla,
                registro.Pulpa,
                registro.Valor_Primer_Brix,
                registro.Hora_Primer_Brix,
                registro.Valor_Brix_Final,
                registro.Hora_Brix_Final,
                registro.Paquete_250_gr,
                registro.Paquete_500_gr,
                registro.Paquete_5000_gr,
            ])

        # Crear un libro de Excel y una hoja de cálculo
        libro = openpyxl.Workbook()
        hoja = libro.active

        # Escribir los encabezados de la tabla
        encabezados = [
            'Bache',
            'Fecha',
            'Gramos de Mora Usados',
            'Gramos de Azúcar',
            'Gramos de Sorbato',
            'Hora Inicio',
            'Primer Hervor',
            'Pausa de enfriado',
            'Despulpado',
            'Última Cocción',
            'Hora Final',
            'Desechos Mora',
            'Semilla',
            'Pulpa',
            'Valor Primer Brix',
            'Hora Primer Brix',
            'Valor Brix Final',
            'Hora Brix Final',
            'Paquete 250 gr',
            'Paquete 500 gr',
            'Paquete 5000 gr',
        ]
        hoja.append(encabezados)

        # Escribir los datos de los registros
        for fila in datos:
            hoja.append(fila)

        # Crear una tabla con estilo en Excel
        tabla = openpyxl.worksheet.table.Table(displayName="TablaRegistros", ref=f"A1:{openpyxl.utils.get_column_letter(len(encabezados))}{len(datos) + 1}")
        hoja.add_table(tabla)

        # Aplicar bordes a las celdas de la tabla
        borde = Border(
                left=Side(border_style="medium", color="000000"),
                right=Side(border_style="medium", color="000000"),
                top=Side(border_style="medium", color="000000"),
                bottom=Side(border_style="medium", color="000000")
        )
        for row in hoja.iter_rows(min_row=2, max_row=len(datos) + 1, min_col=1, max_col=len(encabezados)):
            for cell in row:
                cell.border = borde
        
        color1 = "F6FDFF"  # Azul
        color2 = "BBDEEA"  # Azul Grisaseo

        # Aplicar los colores intercalados a las filas de la tabla
        fill1 = PatternFill(start_color=color1, end_color=color1, fill_type="solid")
        fill2 = PatternFill(start_color=color2, end_color=color2, fill_type="solid")

        for i, row in enumerate(hoja.iter_rows(min_row=2, max_row=len(datos) + 1, min_col=1, max_col=len(encabezados)), start=1):
            if i % 2 == 0:
                for cell in row:
                    cell.fill = fill1
            else:
                for cell in row:
                    cell.fill = fill2

        # Ajustar el ancho de las columnas
        for columna in hoja.columns:
            max_length = 0
            column = columna[0].column_letter  # Obtiene la letra de la columna
            for cell in columna:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.5  # Ajusta el ancho de la columna
            hoja.column_dimensions[column].width = adjusted_width

        # Crear una respuesta HTTP con el archivo de Excel como contenido
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Registros.xlsx'

        # Guardar el libro de Excel en el objeto response
        libro.save(response)

        return response

class BuscarRegistroView(ListView):
    model = Registro
    template_name = 'registros/busqueda.html'
    context_object_name = 'registros'
    paginate_by = 10

    def get_queryset(self):
        query = self.request.GET.get('q')
        if query:
            return Registro.objects.filter(Q(Bache__icontains=query))
        else:
            return Registro.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        return context
    
class RegistroUsuarioView(SuccessMessageMixin, CreateView):
    template_name = 'registros/registro_usuario.html'
    form_class = RegistrodeUsuarioForm
    success_url = reverse_lazy('registros:home')
    success_message = "Usuario %(username)s creado"

    def form_valid(self, form):
        response = super().form_valid(form)
        username = form.cleaned_data['username']
        messages.success(self.request, self.success_message % {'username': username})
        return response
    
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        for field in self.form_class.base_fields.values():
            field.help_text = ''

class LoginView(LoginView):
    template_name = 'registros/login.html'
    success_url = reverse_lazy('registros:home')

class MiLogoutView(LogoutView):
    next_page = reverse_lazy('registros:login')



    # def crear_hoja(registros, nombre_hoja):
#     # Crear libro de excel
#     libro = openpyxl.Workbook()
#     hoja = libro.active
#     hoja.title = nombre_hoja

#     # Escribir encabezados de columna
#     encabezados = ['Bache', 'Fecha', 'Gramos de Mora', 'Gramos de Azucar', 'Gramos de Sorbato', 'Hora Inicio', 'Primer Hervor', 'Pausa de enfriado', 'Despulpado', 'Ultima Coccion', 'Hora Final', 'Desechos Mora', 'Semilla', 'Pulpa', 'Valor Primer Brix', 'Hora Primer Brix', 'Valor Brix Final', 'Hora Brix Final', 'Paquete 250 gr', 'Paquete 500 gr', 'Paquete 5000 gr']
#     for i, encabezado in enumerate(encabezados):
#         celda = hoja.cell(row=1, column=i+1)
#         celda.value = encabezado

#     # Escribir datos de registros
#     for i, registro in enumerate(registros):
#         fila = i+2  # Empezar en la segunda fila
#         hoja.cell(row=fila, column=1, value=registro.Bache)
#         hoja.cell(row=fila, column=2, value=registro.Fecha)
#         hoja.cell(row=fila, column=3, value=registro.Gramos_de_Mora)
#         hoja.cell(row=fila, column=4, value=registro.Gramos_de_Azucar)
#         hoja.cell(row=fila, column=5, value=registro.Gramos_de_Sorbato)
#         hoja.cell(row=fila, column=6, value=registro.Hora_Inicio)
#         hoja.cell(row=fila, column=7, value=registro.Primer_Hervor)
#         hoja.cell(row=fila, column=8, value=registro.Pausa_de_enfriado)
#         hoja.cell(row=fila, column=9, value=registro.Despulpado)
#         hoja.cell(row=fila, column=10, value=registro.Ultima_Coccion)
#         hoja.cell(row=fila, column=11, value=registro.Hora_Final)
#         hoja.cell(row=fila, column=12, value=registro.Desechos_Mora)
#         hoja.cell(row=fila, column=13, value=registro.Semilla)
#         hoja.cell(row=fila, column=14, value=registro.Pulpa)
#         hoja.cell(row=fila, column=15, value=registro.Valor_Primer_Brix)
#         hoja.cell(row=fila, column=16, value=registro.Hora_Primer_Brix)
#         hoja.cell(row=fila, column=17, value=registro.Valor_Brix_Final)
#         hoja.cell(row=fila, column=18, value=registro.Hora_Brix_Final)
#         hoja.cell(row=fila, column=19, value=registro.Paquete_250_gr)
#         hoja.cell(row=fila, column=20, value=registro.Paquete_500_gr)
#         hoja.cell(row=fila, column=21, value=registro.Paquete_5000_gr)

#     return libro

# def crear_archivo_excel(registros):
#     # Crear un libro de Excel y una hoja
#     wb = openpyxl.Workbook()
#     ws = wb.active

#     # Agregar encabezados de columna
#     encabezados = [
#         'Bache',
#         'Fecha',
#         'Gramos de Mora',
#         'Gramos de Azucar',
#         'Gramos de Sorbato',
#         'Hora de Inicio',
#         'Primer Hervor',
#         'Pausa de enfriado',
#         'Despulpado',
#         'Ultima Coccion',
#         'Hora Final',
#         'Desechos Mora',
#         'Semilla',
#         'Pulpa',
#         'Valor Primer Brix',
#         'Hora Primer Brix',
#         'Valor Brix Final',
#         'Hora Brix Final',
#         'Paquete 250 gr',
#         'Paquete 500 gr',
#         'Paquete 5000 gr',
#     ]
#     ws.append(encabezados)

#     # Agregar datos de registros
#     for registro in registros:
#         datos_registro = [
#             registro.Bache,
#             registro.Fecha,
#             registro.Gramos_de_Mora,
#             registro.Gramos_de_Azucar,
#             registro.Gramos_de_Sorbato,
#             registro.Hora_Inicio,
#             registro.Primer_Hervor,
#             registro.Pausa_de_enfriado,
#             registro.Despulpado,
#             registro.Ultima_Coccion,
#             registro.Hora_Final,
#             registro.Desechos_Mora,
#             registro.Semilla,
#             registro.Pulpa,
#             registro.Valor_Primer_Brix,
#             registro.Hora_Primer_Brix,
#             registro.Valor_Brix_Final,
#             registro.Hora_Brix_Final,
#             registro.Paquete_250_gr,
#             registro.Paquete_500_gr,
#             registro.Paquete_5000_gr,
#         ]
#         ws.append(datos_registro)

#     # Crear respuesta HTTP y agregar archivo Excel a ella
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename=registros.xlsx'
#     wb.save(response)

#     return response



